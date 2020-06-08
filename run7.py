
from openpyxl.styles import Font, Alignment, PatternFill, colors


from run8 import  *



from openpyxl.utils import FORMULAE

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))



class report:



    def __init__(self,root):

        self.list_original=create_all_list()[1] # c'est la list des titres de tableau de fichier score....
        self.list_report=create_all_list()[0] # c'est le global des titrs qui se trouve dans  le rapport final
        self.list_report2=create_all_list()[2] # c'est pour regler le dicalage ici ona utiliser list_state_titles2 unitile
        self.list_red=[] # list_red c'est l'esemble des noms qui ont couleur rouge

        self.root = root
        self.root.title('kpi')
        self.interface()


    # lire les donnéés avec pandas

    def interface(self):
        self.frame1 = Frame(self.root, width=340, bg='#0000FF',height=300)
        self.frame1.grid(row=0, column=0, ipady=10, ipadx=10)
        self.frame1.grid_propagate(0)

        self.frame1.grid_rowconfigure(0, weight=3)
        self.frame1.grid_rowconfigure(1, weight=3)
        self.frame1.grid_rowconfigure(2, weight=3)
        self.frame1.grid_rowconfigure(3, weight=2)
        #self.frame1.grid_columnconfigure(0, weight=2)

        self.vfichier1 = StringVar()
        self.vfichier2 = StringVar()
        self.vfichier1.set('')
        self.vfichier2.set('')
        self.chemin = ''
        self.chemin1 = ''

        self.button1 = Button(self.frame1, text="ScoreCard File", command=self.set_fichier1,width=50,
                              height=2, bg='#66B239')
        self.button1.grid(row=0, column=0,columnspan=3, pady=5)

        self.button2 = Button(self.frame1, text='eClem File',command=self.set_fichier2, width=50, height=2,
                              bg='#66B239')
        self.button2.grid(row=1, column=0,columnspan=3, pady=5)

        self.button3 = Button(self.frame1, text='Generate report',command=self.set_emplacement, width=50, height=2,
                                 bg='#66B239')
        self.button3.grid(row=2, column=0, pady=5)
        self.progress_bar = Progressbar(self.frame1, orient='horizontal', length=286, mode='determinate')

    def set_fichier1(self):

        self.FILETYPES = [("text files", "*.xlsm;*.xlsx")]
        self.vfichier1.set(askopenfilename(filetypes=self.FILETYPES))
        if self.vfichier1.get() != '':
            self.button1['bg'] = '#006738'

    def set_fichier2(self):

        self.FILETYPES = [("text files", "*.xlsx")]
        self.vfichier2.set(askopenfilename(filetypes=self.FILETYPES))
        if self.vfichier2.get() != '':
            self.button2['bg'] = '#006738'

    def set_emplacement(self):
        import time
        self.FILETYPES = [("text files", "*.xlsx")]
        chemin1 = (askdirectory())
        date_now = time.strftime('%d%m%Y')
        def go():
            self.progress_bar.grid(row=3, column=0)

            self.progress_bar["value"] = 5
            self.root.update()

            self.red_items()  # c'est une list de range de liste self.list_names_titles qui [1...10]

            self.create_table()


            self.button1['bg'] = '#66B239'
            self.button2['bg'] = '#66B239'
            time.sleep(1)
            self.progress_bar.grid_forget()
            root.update()
            messagebox.showinfo(title=None,message="report successfully created")

            root.destroy()


        if chemin1 != '':
            self.chemin = chemin1 + '/' + 'F' + date_now + '.xlsx'
            my_file=Path( self.chemin)
            if my_file.exists():
                answer=messagebox.askquestion('file exists','The file already exists. Do you want to replace it? ')
                if answer=='yes':
                    go()


                else:None

            else:
                go()



    def read_data(self):
        import copy
        self.data1 = pd.read_excel(self.vfichier1.get(), sheet_name='VehicleScoreCard', skiprows=5, usecols='A : BQ')
        self.data2 = pd.read_excel(self.vfichier2.get(), sheet_name='ECU_Service_SWQA-strategy',usecols='A : BQ')


        self.list_service = list(self.data2['Service'])
        self.list_id = list(self.data2['ID CLEM O52_ASWP'])
        self.list_KPI1_c = list(self.data2['KPI1.c'])
        self.list_KPI1_d = list(self.data2['KPI1.d'])

        self.list_id_net=[]
        self.list_plus=[]
        for i in list(self.data1['ID']):
            if i !='nan':
                self.list_id_net.append(i)




    def insert_KPI(self,rapport,workbook): #inserer les données de KPI qui se trouvr dans le fichiers suivi.. sheet ..stratigie  dans la fonction reand data
        self.progress_bar["value"] = 60
        root.update()

        header_formattxt = Alignment(wrap_text=True)

        list_indice1=['KPI.1c\n'+i for i in list_indice]
        list_indice2=['KPI.1d\n'+i for i in list_indice]

        self.kpi_c=[]
        self.kpi_cv=[] #pour rapport2
        self.kpi_d=[]
        self.kpi_dv=[]
        kpic_instance=[]
        for i in range(len(self.list_report)):# l'objectif c'est trouver lindex de KPI dans le tableau de rapport pour
            # inserer les donné dans la coloune qui a ce index
            x = 6
            kpic_instance = []

            kpicv_instance = []

            kpid_instance = []

            kpidv_instance = []

            if self.list_report[i] in list_indice1:


                for m, n in zip(self.list_KPI1_c, self.list_id):


                    x += 1
                    if n in list(self.data1['ID']):
                        v = list(self.data1['ID']).index(int(n)) + 7
                        k = "Only check rules eCLEM fill"
                        if k in str(m):

                            rapport.cell(v, i + 1).value = "Only check \n" \
                                                           " rules eCLEM \nfill"
                            rapport.cell(v, i + 1).border = thin_border
                            rapport.cell(v, i + 1).alignment = header_formattxt
                            rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                            kpic_instance.append("Only check \n" \
                                                 " rules eCLEM \nfill")

                            kpicv_instance.append(v-7)

                        else:


                            rapport.cell(v, i + 1).value = str(rapport.cell(v, i).value)
                            rapport.cell(v, i + 1).border = thin_border
                            rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)
                            kpic_instance.append(str(rapport.cell(v, i).value))
                            kpicv_instance.append(v-7)

                        rapport.column_dimensions[get_column_letter(i + 1)].width = 15 # ici la dimmention des case de KPI


                self.kpi_c.append(kpic_instance)
                self.kpi_cv.append(kpicv_instance)





            if self.list_report[i] in list_indice2:
                for m, n in zip(self.list_KPI1_d, self.list_id):
                    x += 1
                    if n in list(self.data1['ID']):
                        v = list(self.data1['ID']).index(int(n)) + 7
                        k = "NA"
                        if k in str(m):

                            rapport.cell(v, i + 1).value = '"NA"'
                            rapport.cell(v, i + 1).border = thin_border
                            rapport.cell(v, i + 1).alignment = header_formattxt
                            rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                            kpidv_instance.append(v - 7)
                            kpid_instance.append('"NA"')

                        else:

                            rapport.cell(v, i + 1).value = str(rapport.cell(v, i).value)
                            #rapport.cell(v, i + 1).value = "hello"

                            rapport.cell(v, i + 1).border = thin_border
                            rapport.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)
                            kpidv_instance.append(v - 7)
                            cor='=rapport!'+str(get_column_letter(i))+str(v)

                            kpid_instance.append(cor)
                        rapport.column_dimensions[get_column_letter(i + 1)].width = 15
                self.kpi_d.append(kpid_instance)
                self.kpi_dv.append(kpidv_instance)


        self.state2(self.rapport2_id, workbook, self.kpi_raport2,self.kpi_c,self.kpi_cv,self.kpi_d,self.kpi_dv)

        go = report3(self.data1, self.data2, workbook, self.list_red)
        #w = go.workbook

    def insert_data(self,rapport,workbook):# inserer les donneé de ID name , safty ,.. qui se trouve dans le fichier score
        self.progress_bar["value"] = 15
        root.update()

        self.idd=0

        self.rapport2_id=[]
        ##########################33

        header_formattxt = Alignment(wrap_text=True)

        for i in  range(len(self.list_report)):
            x = 6
            if self.list_report[i] in self.list_original and self.list_report[i] not in [' ','Unnamed','Service']:
                b = []
                for m in list(self.data1[self.list_report[i]]):

                    x += 1
                    if x<len( self.data1[self.list_report[i]])+7:# pour supprimer les case qui contient la valeur nan dans letableau score
                        rapport.cell(x, i + 1).value = m
                        b.append(m)
                        rapport.cell(x, i+1).border = thin_border
                self.rapport2_id.append(b)

            if self.list_report[i]=='Service':# inserer la colone service qui se trouve dans le fichier suivi ...
                b = []
                for m, n in zip(self.list_service, self.list_id):

                    x += 1
                    self.idd += 1
                    if n in list(self.data1['ID']):

                        v = list(self.data1['ID']).index(int(n)) + 7
                        rapport.cell(v, i+1).value = m

                        #b.append(self.list_service[v-1])

                        b.append(v-7)
                        rapport.cell(v, i+1).border = thin_border

                self.rapport2_id.append(b)

        self.state(rapport )
        self.insert_KPI(rapport,workbook)

        #self.state2(self.rapport2_id,workbook,[])


        ####################################################################################################################



        workbook.save(self.chemin)
        workbook.close()




    def create_table(self): #creatio le tableau de rapport inserer les titres

        v=0
        list_jaune=[]
        list_bleu=[]
        self.read_data()

        for i in list_names_titles:
            x=list_indice[v]
            v+=1
            self.list_state_titles = ['Nb "G"', 'Nb "O"', 'Nb "R"', 'Nb "NA"', 'Nb "NE"', '% G\n' + x,
                                      '% O\n' + x,
                                      '% R\n' + x, '% NA\n' + x, '% NE\n' + x, ' ', 'N',
                                      'n',
                                      'KPI.1a\n' + x,
                                      'KPI.1b\n' + x, 'n-Nb"R"', 'V', 'Unnamed', 'KPI.1c\n' + x,
                                      'Nb "G"', 'Nb "O"', 'Nb "R"',
                                      'Unnamed', 'KPI.1d\n' + x,
                                      ]
            list_jaune.extend([self.list_state_titles[i] for i in [13, 14, 18, 23]])
            list_bleu.extend([self.list_state_titles[i] for i in [5, 6, 7, 8, 9]])

        self.my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        self.my_green = openpyxl.styles.colors.Color(rgb='0000FF00')
        self.my_orange = openpyxl.styles.colors.Color(rgb='00FFA500')
        self.my_gray = openpyxl.styles.colors.Color(rgb='00696969')
        self.my_yallow = openpyxl.styles.colors.Color(rgb='00FFFF00')
        my_blue = openpyxl.styles.colors.Color(rgb='001E90FF')
        self.my_black = openpyxl.styles.colors.Color(rgb='00000000')
        header_formatfont = Font(bold=True, )
        header_formattxt = Alignment(wrap_text=True)

        workbook = Workbook()
        self.workboo2=workbook
        rapport = workbook.active

        rapport.title = 'report1'
        ref= 'A6:'+str(get_column_letter(len(self.list_report)))+str(len(self.list_id))


        tab = Table(displayName="Table1", ref=ref)
        # I list out the 4 show-xyz options here for reference
        style = TableStyleInfo(
            #name="TableStyleLight18",
            name="TableStyleLight21",
            #name="TableStyleMedium22",

            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        rapport.add_table(tab)
        x=6

        for i in  range(len(list_report)):
            rapport.cell(x,i+1).value=self.list_report[i]
            rapport.cell(x, i+1).border = thin_border
            if rapport.cell(x,i+1).value in list_head[0] :
                rapport.cell(x,i+1).font = header_formatfont # font gras pour les titres
                rapport.column_dimensions[get_column_letter(i + 1)].width = 15
            if rapport.cell(x,i+1).value in self.list_red:rapport.cell(x,i+1).font = Font(color="FFFF0000")# mette les coulleur rouge
            if rapport.cell(x,i+1).value in list_jaune:
                rapport.cell(x,i+1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)# mette les coulleur jaune
                rapport.column_dimensions[get_column_letter(i + 1)].width = 15
            if rapport.cell(x,i+1).value in list_bleu:
                rapport.cell(x,i+1).fill = PatternFill(patternType='solid', fgColor=my_blue)# mette les coulleur bleu
                rapport.column_dimensions[get_column_letter(i + 1)].width = 15

            rapport.cell(x-1,i+1).fill = PatternFill(patternType='solid', fgColor=self.my_black)
            rapport.cell(x,i+1).alignment = header_formattxt



        self.insert_data(rapport,workbook)




    def red_items(self):
        list_red_index=[]
        wb = openpyxl.load_workbook(filename=self.vfichier1.get())
        sheet_ranges = wb['VehicleScoreCard']
        for i in range(1, len(self.list_original)):
            color_obj = sheet_ranges.cell(row=6, column=i ).font.color
            if color_obj is not None:
                if color_obj.rgb == "FFFF0000": # pour ajouter les titre qui sont on couleur rouge pour les prendre en considiration
                    list_red_index.append(self.list_report.index(self.list_original[i]))

        for i in list_red_index:
            self.list_red .append(self.list_report[i])





    def state2(self, rap2_id, workbook,rap2_kpi_ab,kpi_c,kpi_cv,kpi_d,kpi_dv):#raport2
        self.progress_bar["value"] = 80
        root.update()

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        header_formatfont = Font(bold=True, )
        header_formattxt = Alignment(wrap_text=True)

        rapport = workbook.create_sheet('report2')  # creation le deuxieme rapport


        a = 0



        list_head = [['ID', 'Name', 'Safety', 'MBD', 'SwQA', 'Direction', 'Service'], ]
        list_kpi = ["KPI1.a", "KPI1.b", "KPI1.c", "KPI.1 délivrable status"]

        list_indice = ['Project Management', 'Improvement Management', 'Supporting activities Management',
                       'SW Requirements Management',
                       'SW Architecture & Design Management', 'SW Coding/Modeling Management',
                       'SW Verification Management', 'SW Qualification Management',
                       'SW Safety Management', 'Supplier Management', 'GLOBAL']
        y = 0
        yy=0
        z = 0

        for k in range(len(rap2_id[0])):
            for j in range(7):


                for i in range(5):
                    if i == 0:

                        c1 = get_column_letter(i + 2 + a) + str(j + 1)
                        c2 = get_column_letter(i + 5 + a) + str(j + 1)
                        c = c1 + ":" + c2

                        rapport.merge_cells(c)
                        rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 40


                    else:
                        rapport.cell(j + 1, 1+a ).value = list_head[0][j]


                        rapport.cell(1, 2 + a).value = rap2_id[0][y]
                        rapport.cell(2, 2 + a).value = rap2_id[1][y]
                        rapport.cell(3, 2 + a).value = rap2_id[2][y]
                        rapport.cell(4, 2 + a).value = rap2_id[3][y]
                        rapport.cell(5, 2 + a).value = rap2_id[4][y]
                        rapport.cell(6, 2 + a).value = rap2_id[5][y]
                        rapport.cell(j + 1, 1+ a).border = thin_border

                        if rap2_id[0].index(rap2_id[0][y]) in rap2_id[6] :
                          rapport.cell(7, 2 + a).value = self.list_service[rap2_id[6].index(rap2_id[0].index(rap2_id[0][y]))]




                        rapport.cell(j + 1, i + 1 + a).border = thin_border

                        rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 20
                        #rapport.cell(j + 1,  1 + a).font = Font(color="00FF8C00")
                        #rapport.cell(j + 1,  2 + a).font = Font(color="00FF8C00")
                        rapport.cell(j + 1, 1+a).font = Font(bold=True, color="00FF8C00")
                        rapport.cell(j + 1, 2+a).font = Font(bold=True, color="00FF8C00")



                    rapport.cell(j+1, i + 1 + a).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    rapport.cell(j + 1, i + 1+a).font = Font(bold=True, color="00FF8C00")
                    #rapport.cell(j + 1, i + 1).font = Font(bold=True, )


            y+=1
            x = 0
            x2=0
            x3=0

            for j in range(11):


                for i in range(5):
                    if j == 0:
                        if i < 4:
                            rapport.cell(j + 8, i + 2 + a).value = list_kpi[i]
                            rapport.cell(j + 8, i + 1 + a).border = thin_border
                            #rapport.cell(j + 8, i+2 + a).font = Font(color="00FF8C00")
                            rapport.cell(j + 8, i+2 + a).font = Font(bold=True, color="00FF8C00")

                    if 0<i<3:
                        if len(rap2_kpi_ab) >0:
                            rapport.cell(j + 9, i + 1 + a).value = rap2_kpi_ab[z][x]
                            x += 1
                    if i==3:
                        if len(kpi_c) >0:
                            if rap2_id[0].index(rap2_id[0][yy]) in kpi_cv[x2]:
                                rapport.cell(j+9, i+1 + a).value = kpi_c[x2][ kpi_cv[x2].index(rap2_id[0].index(rap2_id[0][yy]))]
                            x2+=1
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 30
                    if i==4:
                        if len(kpi_d) >0:
                            if rap2_id[0].index(rap2_id[0][yy]) in kpi_dv[x3]:
                                rapport.cell(j+9, i+1 + a).value = kpi_d[x3][ kpi_dv[x3].index(rap2_id[0].index(rap2_id[0][yy]))]
                            x3+=1
                            rapport.column_dimensions[get_column_letter(i + 1 + a)].width = 30

                    rapport.cell(j + 9, 1 + a).value = list_indice[j]
                    rapport.cell(j + 9, i + 1 + a).border = thin_border
                    rapport.cell(j + 9, i + 1+a).font = Font(bold=True,color="00FF8C00" )
                    #rapport.cell(j+9, i + 1 + a).font = Font(color="00FF8C00")
                    rapport.cell(j + 9, i + 1 + a).alignment = Alignment(horizontal='center', vertical='center')


            a += 6
            z+=1
            yy += 1



















        ##########################################################################################################################################################

    def state(self, rapport ):
        self.progress_bar["value"] = 30
        root.update()

        def traiter(args):
            i_itmes = []
            for i in range(len(list_names_titles)):
                i_itmes.append(i)  # pour vérifier est ce que self.valeur est compatible
            w = 6
            a = 0
            self.kpi_raport2 = []
            kpi_instance = []

            for k in range(len(self.list_id_net)):

                w += 1
                s = -1
                kpi_instance = []
                for arg in args:


                    gg = 0
                    oo = 0
                    rr = 0
                    nna = 0
                    nne = 0
                    pper = 0
                    fglobal = '='
                    s += 1
                    arg.sort()

                    for i, j in zip(i_itmes, arg):
                        if i in self.valeur:
                            g = 0
                            o = 0
                            r = 0
                            na = 0
                            ne = 0
                            per = 0

                            for n in self.list_red:
                                if n in list_names_titles[i]:
                                    per += 1
                                    pper += 1
                                    instance = list(self.data1[n])[k]

                                    if str(instance).strip() == 'G':
                                        g += 1
                                        gg += 1
                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_green)

                                    if str(instance).strip() == 'O':
                                        o += 1
                                        oo += 1
                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_orange)
                                    if str(instance).strip() == 'R':
                                        r += 1
                                        rr += 1

                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_red)
                                    if str(instance).strip() == 'nan':
                                        na += 1
                                        nna += 1
                                        rapport.cell(w, self.list_report.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_gray)
                                    if str(instance).strip() == 'NE':
                                        ne += 1
                                        nne += 1

                            ################################lie a tout les pertie qui des titre rouge###############################
                            go = [g, o, r, na, ne]

                            rapport.cell(w, j + 1).value = go[s]

                            fx1 = '=NB.SI(' + str(get_column_letter(j - 5)) + str(w) + ':' + str(
                                get_column_letter(j)) + str(w) + ',"O")'
                            # fx2='=LEN( '+str(get_column_letter(j -5))+str(w)+')-LEN(SUBSTITUTE('+ str(get_column_letter(j -5))+str(w)+',"z",""))'
                            # rapport.cell(w, j + 1).value = fx2

                            # rapport.cell(w, j + 1).value = '= NB.SI.ENS(I7:N7;"O")'

                            # rapport.cell(w, j + 6).value = str(round(100*go[s]/per,2))+"%"
                            fx6 = '=' + str(get_column_letter(j + 1)) + str(w) + '/' + str(per)

                            # rapport.cell(w, j + 6).value = fx6
                            rapport.cell(w, j + 6).value = str(round((100 * go[s] / per), 2)) + "%"
                            rapport.cell(w, j + 1).border = thin_border
                            rapport.cell(w, j + 6).border = thin_border
                            if s == len(go) - 1:

                                if a < len(list_indice):
                                    rapport.cell(4, j + 15).value = list_indice[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport.cell(4, j - 3).value = list_indice[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    a += 1

                                rapport.cell(w, j + 8).border = thin_border
                                rapport.cell(w, j + 9).border = thin_border
                                rapport.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_gray)
                                rapport.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                rapport.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                               fgColor=self.my_black)  # le vide noir entre les changement
                                rapport.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                           fgColor=self.my_black)  # le vide noir entre les changement
                                rapport.cell(w, j + 8).value = per  # N
                                rapport.cell(w, j + 9).value = per  # n

                                rapport.cell(w, j + 10).value = str(
                                    per / per * 100) + '%'  # KPI.1a Project Management###########################################################

                                rapport.cell(w, j + 11).value = str(round(((go[0] * 1 + go[1] * 0.5 + go[2] * 0) / per * 100), 2)) + '%'  # KPI.1bProject Management
                                kpi_instance.append(str(per / per * 100) + '%')
                                kpi_instance.append(str(round(((go[0] * 1 + go[1] * 0.5 + go[2] * 0) / per * 100), 2)) + '%' )# pour l'envoyer au rapport2

                                rapport.cell(w, j + 12).value = sum(go) - go[2]  # n-Nb"R"
                                rapport.cell(w, j + 13).value = go[0]  # V
                                if go[2] != 4:
                                    try:
                                        rapport.cell(w, j + 14).value = str(
                                            round(((100 / (sum(go) - go[2])) * go[0]), 2)) + '%'
                                    except ZeroDivisionError as error:
                                        rapport.cell(w, j + 14).value ='0,00%'


                                else:
                                    rapport.cell(w, j + 14).value = '0' + '%'
                                rapport.cell(w, j + 16).value = 0
                                rapport.cell(w, j + 17).value = 0
                                rapport.cell(w, j + 18).value = 0
                                # fonction pour automatiquer de calcule
                                fx19 = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)*100/' + str(
                                    get_column_letter(j + 13)) + str(w)  # pour  automatiser
                                fxglo = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                    get_column_letter(j + 13)) + str(w) + '+'  # pour  automatiser
                                fglobal += fxglo
                                # rapport.cell(w, j + 19).value = '= (BH7*1 + BI7*0,5 + BJ7*0) /BE7'
                                # rapport.cell(w, j + 19).value = fx19
                                rapport.cell(w,
                                             j + 19).value = '=IF(' + fx19 + ' > 0," "' + '& ' + fx19 + ' & "%",' + fx19 + ' & "%")'

                        ########################################partie global lie a mist gooo############################################
                        elif i == i_itmes[-1]:  # i est l index de  dernier parier de GLOBAL
                            goo = [gg, oo, rr, nna, nne]
                            rapport.cell(w, j + 1).value = goo[s]
                            fx6 = '=' + str(get_column_letter(j + 1)) + str(w) + '/' + str(pper) + '*100 '

                            rapport.cell(w, j + 6).value = fx6
                            rapport.cell(w, j + 1).border = thin_border
                            rapport.cell(w, j + 6).border = thin_border
                            if s == len(goo) - 1:

                                if a < len(list_indice):
                                    rapport.cell(4, j + 15).value = list_indice[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport.cell(4, j - 3).value = list_indice[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    a += 1

                                rapport.cell(w, j + 8).border = thin_border
                                rapport.cell(w, j + 9).border = thin_border
                                rapport.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_gray)
                                rapport.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                rapport.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                               fgColor=self.my_black)  # le vide noir entre les changement
                                rapport.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                           fgColor=self.my_black)  # le vide noir entre les changement
                                rapport.cell(w, j + 8).value = pper  # N
                                rapport.cell(w, j + 9).value = pper  # n

                                rapport.cell(w, j + 10).value = str(
                                    pper / pper * 100) + '%'  # KPI.1a Project Management
                                rapport.cell(w, j + 11).value = str(
                                    round(((goo[0] * 1 + goo[1] * 0.5 + goo[2] * 0) / pper * 100),
                                          2)) + '%'  # KPI.1bProject Management

                                kpi_instance.append(str(pper / pper * 100) + '%')
                                kpi_instance.append(str(round(((goo[0] * 1 + goo[1] * 0.5 + goo[2] * 0) / pper * 100), 2)) + '%' )  # pour l'envoyer au rapport2
                                rapport.cell(w, j + 12).value = sum(goo) - goo[2]  # n-Nb"R"
                                rapport.cell(w, j + 13).value = goo[0]  # V
                                if goo[2] != 4:
                                    rapport.cell(w, j + 14).value = str(
                                        round(((100 / (sum(goo) - goo[2])) * goo[0]), 2)) + '%'
                                else:
                                    rapport.cell(w, j + 14).value = '0' + '%'
                                rapport.cell(w, j + 16).value = 0
                                rapport.cell(w, j + 17).value = 0
                                rapport.cell(w, j + 18).value = 0
                                # fonction pour automatiquer de calcule
                                fx19 = '=(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                    get_column_letter(j + 13)) + str(
                                    w)  # pour  automatiser
                                # rapport.cell(w, j + 19).value = '= (BH7*1 + BI7*0,5 + BJ7*0) /BE7'
                                # =SI((B2 - A2) / A2 % > 0;"+ " & (B2 - A2) / A2 % & " %";(B2 - A2) / A2 % & " %")
                                # ="=SI("+fxglo2 +" > 0;"+ " & "+fxglo2+" &  '%';"+fxglo2+"& " '%'")

                                fxglo2 = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*1/5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                    get_column_letter(j + 13)) + str(
                                    w)
                                fxglo = '=(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                    get_column_letter(j + 13)) + str(
                                    w)  # pour  automatiser
                                # rapport.cell(w, j + 19).value = fglobal

                                # rapport.cell(w, j + 19).value =  fxglo
                                # rapport.cell(w, j + 19).value =  '=IF('+fxglo2 +' > 0;"+"'+ '& '+fxglo2+' & "%";'+fxglo2+' & "%")'
                                # pour marche mette ; ou lieu de  ,
                                rapport.cell(w,
                                             j + 19).value = '=IF(' + fxglo2 + ' > 0," "' + '& ' + fxglo2 + ' & "%",' + fxglo2 + ' & "%")'
                                # rapport.cell(w, j + 19).value = 'IF(+'+str(fxglo2)+'  > 0,'+""+str(fxglo2)+"'%'"+',"False")'



                        ###############################################parite lie a les elemnt qui ont aps aucun titre rouge
                        else:

                            go = [1, 2, 3, 4, 5]
                            rapport.cell(w, j + 1).value = 0
                            rapport.cell(w, j + 6).value = "0%"
                            rapport.cell(w, j + 1).border = thin_border
                            rapport.cell(w, j + 6).border = thin_border
                            if s == len(go) - 1:
                                if a < len(list_indice):
                                    rapport.cell(4, j + 15).value = list_indice[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport.cell(4, j - 3).value = list_indice[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    a += 1
                                rapport.cell(w, j + 8).border = thin_border
                                rapport.cell(w, j + 9).border = thin_border
                                rapport.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                              fgColor=self.my_gray)  # c'est pour siparation en gray '
                                rapport.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                rapport.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                               fgColor=self.my_black)
                                rapport.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                           fgColor=self.my_black)
                                rapport.cell(w, j + 8).value = 0  # N
                                rapport.cell(w, j + 9).value = 0  # n
                                rapport.cell(w, j + 10).value = '0%'  # KPI.1a Project Management
                                rapport.cell(w, j + 11).value = '0%'
                                kpi_instance.append('0%')# pour l'envoyer
                                kpi_instance.append('0%')

                                rapport.cell(w, j + 12).value = 0  # n-Nb"R"
                                rapport.cell(w, j + 13).value = 0  # V
                                if go[2] != 4:
                                    rapport.cell(w, j + 14).value = '0%'
                                else:
                                    rapport.cell(w, j + 14).value = '0' + '%'

                self.kpi_raport2.append(kpi_instance)



        g_index = [i for i, x in enumerate(self.list_report) if x == self.list_state_titles[
            0]]  # get l'index des case qui va prend la somme des couleur greende list fix
        o_index = [i for i, x in enumerate(self.list_report) if x == self.list_state_titles[1]]
        r_index = [i for i, x in enumerate(self.list_report) if x == self.list_state_titles[2]]
        na_index = [i for i, x in enumerate(self.list_report) if x == self.list_state_titles[3]]
        ne_index = [i for i, x in enumerate(self.list_report) if x == self.list_state_titles[4]]
        list_c = [g_index, o_index, r_index, na_index, ne_index]

        self.valeur = []
        self.list_c2 = []  # list des index des coloone de state des couleur

        for i in self.list_report:
            for j in list_names_titles:
                if i in j and i in self.list_red:
                    if list_names_titles.index(j) not in self.valeur:
                        self.valeur.append(list_names_titles.index(j))
                    for p in list_c:
                        self.z = []
                        if p == list_c[3] or p == list_c[4]:
                            for h in range(len(p)):
                                if p[h] not in self.z:
                                    self.z.append(p[h])
                            if self.z not in self.list_c2:
                                self.list_c2.append(self.z)
                        else:
                            for h in range(len(p)):
                                if h % 2 == 0:
                                    if p[h] not in self.z:
                                        self.z.append(p[h])
                            if self.z not in self.list_c2:
                                self.list_c2.append(self.z)

                        # for h in range(len(r_index)) :
                        # if h %2==0:
                        # if r_index[h] not in self.f:
                        #  self.f.append(r_index[h])

        w = 6
        traiter(self.list_c2)



if __name__== "__main__":
    root = Tk()


    report(root)
    root.mainloop()
