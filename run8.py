from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename,askdirectory
from tkinter import messagebox

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, colors

from openpyxl.styles import Font, Alignment, PatternFill, colors ,Color
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from tkinter.ttk import Progressbar
from pathlib import Path

from openpyxl.utils import FORMULAE

import copy
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))



list_indice=['Project Management','Improvement Management','Supporting activities Management','SW Requirements Management',
             'SW Architecture & Design Management','SW Coding/Modeling Management','SW Verification Management','SW Qualification Management',
             'SW Safety Management','Supplier Management','GLOBAL']

#les titrs des paritie de rapport final

list_indice3=['PIM.3 Process Improvement','MAN.3 Project Management','MAN.5 Risk Management','MAN.6 Measurement','ACQ.3 Contract Agreement','ACQ.4 Supplier Monitoring',
              'ACQ.3 Contract Agreement','SUP.1 Quality Assurance','SUP.8 Configuration Management','SUP.9 Problem Resolution Management','SUP.10 Change Request Management',
              'All SWE.X','SWE.2 Software Architectural Design','SWE.3 Software Detailed Design and Unit Construction','SWE.6 Software Qualification Test',
              'SPL.2 Product Release','SWE.1 Software Requirements Analysis','Specific Safety','SWE.4 Software Unit Verification','GLOBAL']

list_names_titles3=[['MAN.I_01'],['MAN.P_01','MAN.P_01a','MAN.P_02','MAN.P_03',],['MAN.P_04','MAN.P_05',],['MAN.P_06'],['MAN.S_01','MAN.S_02','MAN.S_03','MAN.S_04','MAN.S_05',],
                    ['MAN.S_06','MAN.S_07','MAN.S_08',],['MAN.S_09'],['SUP_01','SUP_01a','SUP_02',],['SUP_04','SUP_05','SUP_06','SUP_07',],['SUP_08'],
                    ['SUP_09','SUP_10','SUP_10a','SUP_11',],['SWE_01','SWE_04',],['SWE.2_01','SWE.2_02',],['SWE.3_01','SWE.3_02',],
                    ['SWE.6_01','SWE.6_02','SWE.6_03','SWE.6_04','SWE.6_05','SWE.6_06',],['SWE_02','SWE_03',],['SWE.1_01','SWE.1_02','SWE.1_02a','SWE.1_03',],
                    ['SAF_01','SAF_02','SAF_03',],['SWE.4_01','SWE.4_02','SWE.4_03','SWE.4_10',],[]]

list_state_titles2 = ['Nb "G"', 'Nb "O"', 'Nb "R"', 'Nb "NA"', 'Nb "NE"', '% G\nProject Management',
                     '% O\nProject Management',
                     '% R\nProject Management', '% NA\nProject Management', '% NE\nProject Management', ' ', 'N',
                     'n',
                     'KPI.1a\nProject Management',
                     'KPI.1b\nProject Management', 'n-Nb"R"', 'V', 'Unnamed', 'KPI.1c\nProject Management',
                     'Nb "G"', 'Nb "O"', 'Nb "R"','Nb "NA"', 'Nb "NE"',
                     'Unnamed', 'KPI.1d\nProject Management',
                     ]  # length=27 # est ajouter pour regler le dicalage par ce que la fonctio inumirate touve un problem si il ' a
                         # 'Nb "R"', 'Nb "NA"', 'Nb "NE" deux fois on l'ajoute pour regler le dicalage dans liste


list_head = [['ID', 'Name', 'Safety', 'MBD', 'SwQA', 'Direction', 'Service'], ] # length=7 les titres de présentaion de rapport
list_names_titles = [
    ['MAN.P_01a', 'MAN.P_01', 'MAN.P_02', 'MAN.P_03', 'MAN.P_04', 'MAN.P_05', 'MAN.P_06'],
    ['MAN.I_01', ],
    ['SUP_01a', 'SUP_01', 'SUP_02', 'SUP_04', 'SUP_05', 'SUP_06', 'SUP_07', 'SUP_08', 'SUP_09', 'SUP_10a', 'SUP_10',
     'SUP_11', 'SWE_01', 'SWE_04', ],
    ['SWE.1_01', 'SWE.1_02a', 'SWE.1_02', 'SWE.1_03', ],
    ['SWE.2_01', 'SWE.2_02', ],
    ['SWE.3_01', 'SWE.3_02', ],
    ['SWE.4_01', 'SWE.4_02', 'SWE.4_03', 'SWE.4_10', ],
    ['SWE.6_01', 'SWE.6_02', 'SWE.6_03', 'SWE.6_04', 'SWE.6_05', 'SWE.6_06', 'SWE_02', 'SWE_03', ],
    ['SAF_01', 'SAF_02', 'SAF_03', ],
    ['MAN.S_01', 'MAN.S_02', 'MAN.S_03', 'MAN.S_04', 'MAN.S_05', 'MAN.S_06', 'MAN.S_07', 'MAN.S_08',
     'MAN.S_09', ],[] ]# length=10 groupe of names names





def create_all_list():


    a = []
    v=0
    for i in list_names_titles:# creer self.list_report
        x=list_indice[v]
        list_state_titles = ['Nb "G"', 'Nb "O"', 'Nb "R"', 'Nb "NA"', 'Nb "NE"', '% G\n' + x,
                                  '% O\n' + x,
                                  '% R\n' + x, '% NA\n' + x, '% NE\n' + x, ' ', 'N',
                                  'n',
                                  'KPI.1a\n' + x,
                                  'KPI.1b\n' + x, 'n-Nb"R"', 'V', 'Unnamed', 'KPI.1c\n' + x,
                                  'Nb "G"', 'Nb "O"', 'Nb "R"',
                                  'Unnamed', 'KPI.1d\n' + x,
                                  ]
        v+=1
        instance = []
        instance = i + list_state_titles
        for j in instance:
            a.append(j)

        a.append(" ")
    a=list_head[0]+a
    b = []
    for i in list_names_titles:# creer la list self.list_original
        for j in i:
            b.append(j)
        b.append(" ")
    b=list_head[0]+b
    c=[]

    for i in list_names_titles:#creer self.list_report2 qui va regler le dicalage dans la liste
        instance = []
        instance = i + list_state_titles2
        for j in instance:
            c.append(j)

        c.append(" ")
    c = list_head[0] + c

    d = []
    v = 0
    for i in list_names_titles3:  # creer self.list_report
        x = list_indice3[v]
        list_state_titles = ['Nb "G"', 'Nb "O"', 'Nb "R"', 'Nb "NA"', 'Nb "NE"', '% G\n' + x,
                             '% O\n' + x,
                             '% R\n' + x, '% NA\n' + x, '% NE\n' + x, ' ', 'N',
                             'n',
                             'KPI.1a\n' + x,
                             'KPI.1b\n' + x, 'n-Nb"R"', 'V', 'Unnamed', 'KPI.1c\n' + x,
                             'Nb "G"', 'Nb "O"', 'Nb "R"',
                             'Unnamed', 'KPI.1d\n' + x,
                             ]
        v += 1
        instance = []
        instance = i + list_state_titles
        for j in instance:
            d.append(j)

        d.append(" ")
    d = list_head[0] + d

    return a, b ,c ,d






list_original=create_all_list()[1] # c'est la list des titres de tableau de fichier score....
list_report=create_all_list()[0] # c'est le global des titrs qui se trouve dans  le rapport final
list_report2=create_all_list()[2] # c'est pour regler le dicalage ici ona utiliser list_state_titles2 unitile
 # list_red c'est l'esemble des noms qui ont couleur rouge
list_report3=create_all_list()[3]


class report3:



    def __init__(self,data1,data2,workbook,list_red ):
        self.list_red=list_red

        self.workbook=workbook
        self.data1=data1
        self.data2=data2
        self.list_original=create_all_list()[1] # c'est la list des titres de tableau de fichier score....

        self.list_report2=create_all_list()[2] # c'est pour regler le dicalage ici ona utiliser list_state_titles2 unitile
        # list_red c'est l'esemble des noms qui ont couleur rouge
        self.list_report3 = create_all_list()[3]
        self.create_table()



    # lire les donnéés avec pandas



    def read_data(self):


        self.list_service = list(self.data2['Service'])
        self.list_id = list(self.data2['ID CLEM O52_ASWP'])
        self.list_KPI1_c = list(self.data2['KPI1.c'])
        self.list_KPI1_d = list(self.data2['KPI1.d'])

        self.list_id_net=[]
        self.list_plus=[]
        for i in list(self.data1['ID']):
            if i !='nan':
                self.list_id_net.append(i)




    def insert_KPI(self, rapport3, workbook): #inserer les données de KPI qui se trouvr dans le fichiers suivi.. sheet ..stratigie  dans la fonction reand data


        header_formattxt = Alignment(wrap_text=True)

        list_indice1=['KPI.1c\n'+i for i in list_indice3]
        list_indice2=['KPI.1d\n'+i for i in list_indice3]

        self.kpi_c=[]
        self.kpi_cv=[] #pour rapport2
        self.kpi_d=[]
        self.kpi_dv=[]
        kpic_instance=[]
        for i in range(len(self.list_report3)):# l'objectif c'est trouver lindex de KPI dans le tableau de rapport pour
            # inserer les donné dans la coloune qui a ce index
            x = 6
            kpic_instance = []

            kpicv_instance = []

            kpid_instance = []

            kpidv_instance = []

            if self.list_report3[i] in list_indice1:


                for m, n in zip(self.list_KPI1_c, self.list_id):


                    x += 1
                    if n in list(self.data1['ID']):
                        v = list(self.data1['ID']).index(int(n)) + 7
                        k = "Only check rules eCLEM fill"
                        if k in str(m):

                            rapport3.cell(v, i + 1).value = "Only check \n" \
                                                           " rules eCLEM \nfill"
                            rapport3.cell(v, i + 1).border = thin_border
                            rapport3.cell(v, i + 1).alignment = header_formattxt
                            rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                            kpic_instance.append("Only check \n" \
                                                 " rules eCLEM \nfill")

                            kpicv_instance.append(v-7)

                        else:


                            rapport3.cell(v, i + 1).value = str(rapport3.cell(v, i).value)
                            rapport3.cell(v, i + 1).border = thin_border
                            rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)
                            kpic_instance.append(str(rapport3.cell(v, i).value))
                            kpicv_instance.append(v-7)

                        rapport3.column_dimensions[get_column_letter(i + 1)].width = 15 # ici la dimmention des case de KPI


                self.kpi_c.append(kpic_instance)
                self.kpi_cv.append(kpicv_instance)





            if self.list_report3[i] in list_indice2:
                for m, n in zip(self.list_KPI1_d, self.list_id):
                    x += 1
                    if n in list(self.data1['ID']):
                        v = list(self.data1['ID']).index(int(n)) + 7
                        k = "NA"
                        if k in str(m):

                            rapport3.cell(v, i + 1).value = '"NA"'
                            rapport3.cell(v, i + 1).border = thin_border
                            rapport3.cell(v, i + 1).alignment = header_formattxt
                            rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                            kpidv_instance.append(v - 7)
                            kpid_instance.append('"NA"')

                        else:

                            rapport3.cell(v, i + 1).value = str(rapport3.cell(v, i).value)
                            #rapport.cell(v, i + 1).value = "hello"

                            rapport3.cell(v, i + 1).border = thin_border
                            rapport3.cell(v, i + 1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)
                            kpidv_instance.append(v - 7)
                            cor='=rapport!'+str(get_column_letter(i))+str(v)

                            kpid_instance.append(cor)
                        rapport3.column_dimensions[get_column_letter(i + 1)].width = 15
                self.kpi_d.append(kpid_instance)
                self.kpi_dv.append(kpidv_instance)



    def insert_data(self,rapport3,workbook):# inserer les donneé de ID name , safty ,.. qui se trouve dans le fichier score

        self.idd=0

        self.rapport2_id=[]
        ##########################33

        header_formattxt = Alignment(wrap_text=True)

        for i in  range(len(self.list_report3)):
            x = 6
            if self.list_report3[i] in self.list_original and self.list_report3[i] not in [' ','Unnamed','Service']:
                b = []
                for m in list(self.data1[self.list_report3[i]]):

                    x += 1
                    if x<len( self.data1[self.list_report3[i]])+7:# pour supprimer les case qui contient la valeur nan dans letableau score
                        rapport3.cell(x, i + 1).value = m
                        b.append(m)
                        rapport3.cell(x, i+1).border = thin_border
                self.rapport2_id.append(b)

            if self.list_report3[i]=='Service':# inserer la colone service qui se trouve dans le fichier suivi ...
                b = []
                for m, n in zip(self.list_service, self.list_id):

                    x += 1
                    self.idd += 1
                    if n in list(self.data1['ID']):

                        v = list(self.data1['ID']).index(int(n)) + 7
                        rapport3.cell(v, i+1).value = m

                        #b.append(self.list_service[v-1])

                        b.append(v-7)
                        rapport3.cell(v, i+1).border = thin_border

                self.rapport2_id.append(b)

        self.state(rapport3 )
        self.insert_KPI(rapport3,self.workbook)

        #self.state2(self.rapport2_id,workbook,[])


        ####################################################################################################################







    def create_table(self): #creatio le tableau de rapport3 inserer les titres

        v=0
        list_jaune=[]
        list_bleu=[]
        self.read_data()

        for i in list_names_titles3:
            x=list_indice3[v]
            v+=1
            self.list_state_titles = ['Nb "G"', 'Nb "O"', 'Nb "R"', 'Nb "NA"', 'Nb "NE"', '% G\n' + x,
                                      '% O\n' + x,
                                      '% R\n' + x, '% NA\n' + x, '% NE\n' + x, ' ', 'N',
                                      'n',
                                      'KPI.1a\n' + x,
                                      'KPI.1b\n' + x, 'n-Nb"R"', 'V', 'Unnamed', 'KPI.1c\n' + x,
                                      'Nb "G"', 'Nb "O"', 'Nb "R"',
                                      'Unnamed', 'KPI.1d\n' + x,
                                      ]
            list_jaune.extend([self.list_state_titles[i] for i in [13, 14, 18, 23]])
            list_bleu.extend([self.list_state_titles[i] for i in [5, 6, 7, 8, 9]])

        self.my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        self.my_green = openpyxl.styles.colors.Color(rgb='0000FF00')
        self.my_orange = openpyxl.styles.colors.Color(rgb='00FFA500')
        self.my_gray = openpyxl.styles.colors.Color(rgb='00696969')
        self.my_yallow = openpyxl.styles.colors.Color(rgb='00FFFF00')
        my_blue = openpyxl.styles.colors.Color(rgb='001E90FF')
        self.my_black = openpyxl.styles.colors.Color(rgb='00000000')
        header_formatfont = Font(bold=True, )
        header_formattxt = Alignment(wrap_text=True)


        rapport3 = self.workbook.create_sheet('report3')

        rapport3.title = 'report3'
        ref= 'A6:'+str(get_column_letter(len(self.list_report3)))+str(len(self.list_id))


        tab = Table(displayName="Table1", ref=ref)
        # I list out the 4 show-xyz options here for reference
        style = TableStyleInfo(
            #name="TableStyleLight18",
            name="TableStyleLight21",
            #name="TableStyleMedium22",

            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        rapport3.add_table(tab)
        x=6

        for i in  range(len(list_report3)):
            rapport3.cell(x,i+1).value=self.list_report3[i]
            rapport3.cell(x, i+1).border = thin_border
            if rapport3.cell(x,i+1).value in list_head[0] :
                rapport3.cell(x,i+1).font = header_formatfont # font gras pour les titres
                rapport3.column_dimensions[get_column_letter(i + 1)].width = 15
            if rapport3.cell(x,i+1).value in self.list_red:rapport3.cell(x,i+1).font = Font(color="FFFF0000")# mette les coulleur rouge
            if rapport3.cell(x,i+1).value in list_jaune:
                rapport3.cell(x,i+1).fill = PatternFill(patternType='solid', fgColor=self.my_yallow)# mette les coulleur jaune
                rapport3.column_dimensions[get_column_letter(i + 1)].width = 15
            if rapport3.cell(x,i+1).value in list_bleu:
                rapport3.cell(x,i+1).fill = PatternFill(patternType='solid', fgColor=my_blue)# mette les coulleur bleu
                rapport3.column_dimensions[get_column_letter(i + 1)].width = 15

            rapport3.cell(x-1,i+1).fill = PatternFill(patternType='solid', fgColor=self.my_black)
            rapport3.cell(x,i+1).alignment = header_formattxt

        self.insert_data(rapport3,self.workbook)










        ##########################################################################################################################################################

    def state(self, rapport3):


        def traiter(args):

            i_itmes = []
            for i in range(len(list_names_titles3)):
                i_itmes.append(i)  # pour vérifier est ce que self.valeur est compatible
            w = 6
            a = 0
            self.kpi_raport2 = []
            kpi_instance = []

            for k in range(len(self.list_id_net)):


                w += 1
                s = -1
                kpi_instance = []
                for arg in args:



                    gg = 0
                    oo = 0
                    rr = 0
                    nna = 0
                    nne = 0
                    pper = 0
                    fglobal = '='
                    s += 1
                    arg.sort()

                    for i, j in zip(i_itmes, arg):


                        if i in self.valeur:
                            g = 0
                            o = 0
                            r = 0
                            na = 0
                            ne = 0
                            per = 0

                            for n in self.list_red:
                                if n in list_names_titles3[i]:
                                    per += 1
                                    pper += 1
                                    instance = list(self.data1[n])[k]

                                    if str(instance).strip() == 'G':
                                        g += 1
                                        gg += 1
                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_green)

                                    if str(instance).strip() == 'O':
                                        o += 1
                                        oo += 1
                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_orange)
                                    if str(instance).strip() == 'R':
                                        r += 1
                                        rr += 1

                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_red)
                                    if str(instance).strip() == 'nan':
                                        na += 1
                                        nna += 1
                                        rapport3.cell(w, self.list_report3.index(n) + 1).fill = PatternFill(
                                            patternType='solid', fgColor=self.my_gray)
                                    if str(instance).strip() == 'NE':
                                        ne += 1
                                        nne += 1

                            ################################lie a tout les pertie qui des titre rouge###############################
                            go = [g, o, r, na, ne]

                            rapport3.cell(w, j + 1).value = go[s]

                            fx1 = '=NB.SI(' + str(get_column_letter(j - 5)) + str(w) + ':' + str(
                                get_column_letter(j)) + str(w) + ',"O")'
                            # fx2='=LEN( '+str(get_column_letter(j -5))+str(w)+')-LEN(SUBSTITUTE('+ str(get_column_letter(j -5))+str(w)+',"z",""))'
                            # rapport.cell(w, j + 1).value = fx2

                            # rapport.cell(w, j + 1).value = '= NB.SI.ENS(I7:N7;"O")'

                            # rapport.cell(w, j + 6).value = str(round(100*go[s]/per,2))+"%"
                            fx6 = '=' + str(get_column_letter(j + 1)) + str(w) + '/' + str(per)

                            # rapport.cell(w, j + 6).value = fx6
                            rapport3.cell(w, j + 6).value = str(round((100 * go[s] / per), 2)) + "%"
                            rapport3.cell(w, j + 1).border = thin_border
                            rapport3.cell(w, j + 6).border = thin_border
                            if s == len(go) - 1:

                                if a < len(list_indice3):
                                    rapport3.cell(4, j + 15).value = list_indice3[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport3.cell(4, j - 3).value = list_indice3[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    a += 1

                                rapport3.cell(w, j + 8).border = thin_border
                                rapport3.cell(w, j + 9).border = thin_border
                                rapport3.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                               fgColor=self.my_gray)
                                rapport3.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                rapport3.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                                fgColor=self.my_black)  # le vide noir entre les changement
                                rapport3.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                            fgColor=self.my_black)  # le vide noir entre les changement
                                rapport3.cell(w, j + 8).value = per  # N
                                rapport3.cell(w, j + 9).value = per  # n

                                rapport3.cell(w, j + 10).value = str(
                                    per / per * 100) + '%'  # KPI.1a Project Management###########################################################

                                rapport3.cell(w, j + 11).value = str(round(((go[0] * 1 + go[1] * 0.5 + go[2] * 0) / per * 100), 2)) + '%'  # KPI.1bProject Management
                                kpi_instance.append(str(per / per * 100) + '%')
                                kpi_instance.append(str(round(((go[0] * 1 + go[1] * 0.5 + go[2] * 0) / per * 100), 2)) + '%' )# pour l'envoyer au rapport2

                                rapport3.cell(w, j + 12).value = sum(go) - go[2]  # n-Nb"R"
                                rapport3.cell(w, j + 13).value = go[0]  # V
                                if go[2] != 4:
                                    try:
                                        rapport3.cell(w, j + 14).value = str(
                                            round(((100 / (sum(go) - go[2])) * go[0]), 2)) + '%'
                                    except ZeroDivisionError as error:
                                        rapport3.cell(w, j + 14).value ='0,00%'

                                else:
                                    rapport3.cell(w, j + 14).value = '0' + '%'
                                rapport3.cell(w, j + 16).value = 0
                                rapport3.cell(w, j + 17).value = 0
                                rapport3.cell(w, j + 18).value = 0
                                # fonction pour automatiquer de calcule
                                fx19 = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)*100/' + str(
                                    get_column_letter(j + 13)) + str(w)  # pour  automatiser
                                fxglo = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                    get_column_letter(j + 13)) + str(w) + '+'  # pour  automatiser
                                fglobal += fxglo
                                # rapport.cell(w, j + 19).value = '= (BH7*1 + BI7*0,5 + BJ7*0) /BE7'
                                # rapport.cell(w, j + 19).value = fx19
                                rapport3.cell(w,
                                              j + 19).value = '=IF(' + fx19 + ' > 0," "' + '& ' + fx19 + ' & "%",' + fx19 + ' & "%")'

                        ########################################partie global lie a mist gooo############################################
                        elif i == i_itmes[-1]:  # i est l index de  dernier parier de GLOBAL
                            goo = [gg, oo, rr, nna, nne]
                            rapport3.cell(w, j + 1).value = goo[s]
                            fx6 = '=' + str(get_column_letter(j + 1)) + str(w) + '/' + str(pper) + '*100 '

                            rapport3.cell(w, j + 6).value = fx6
                            rapport3.cell(w, j + 1).border = thin_border
                            rapport3.cell(w, j + 6).border = thin_border
                            if s == len(goo) - 1:

                                if a < len(list_indice3):
                                    rapport3.cell(4, j + 15).value = list_indice3[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport3.cell(4, j - 3).value = list_indice3[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    a += 1

                                rapport3.cell(w, j + 8).border = thin_border
                                rapport3.cell(w, j + 9).border = thin_border
                                rapport3.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                               fgColor=self.my_gray)
                                rapport3.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                rapport3.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                                fgColor=self.my_black)  # le vide noir entre les changement
                                rapport3.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                            fgColor=self.my_black)  # le vide noir entre les changement
                                rapport3.cell(w, j + 8).value = pper  # N
                                rapport3.cell(w, j + 9).value = pper  # n

                                rapport3.cell(w, j + 10).value = str(
                                    pper / pper * 100) + '%'  # KPI.1a Project Management
                                rapport3.cell(w, j + 11).value = str(
                                    round(((goo[0] * 1 + goo[1] * 0.5 + goo[2] * 0) / pper * 100),
                                          2)) + '%'  # KPI.1bProject Management

                                kpi_instance.append(str(pper / pper * 100) + '%')
                                kpi_instance.append(str(round(((goo[0] * 1 + goo[1] * 0.5 + goo[2] * 0) / pper * 100), 2)) + '%' )  # pour l'envoyer au rapport2
                                rapport3.cell(w, j + 12).value = sum(goo) - goo[2]  # n-Nb"R"
                                rapport3.cell(w, j + 13).value = goo[0]  # V
                                if goo[2] != 4:
                                    rapport3.cell(w, j + 14).value = str(
                                        round(((100 / (sum(goo) - goo[2])) * goo[0]), 2)) + '%'
                                else:
                                    rapport3.cell(w, j + 14).value = '0' + '%'
                                rapport3.cell(w, j + 16).value = 0
                                rapport3.cell(w, j + 17).value = 0
                                rapport3.cell(w, j + 18).value = 0
                                # fonction pour automatiquer de calcule
                                fx19 = '=(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                    get_column_letter(j + 13)) + str(
                                    w)  # pour  automatiser
                                # rapport.cell(w, j + 19).value = '= (BH7*1 + BI7*0,5 + BJ7*0) /BE7'
                                # =SI((B2 - A2) / A2 % > 0;"+ " & (B2 - A2) / A2 % & " %";(B2 - A2) / A2 % & " %")
                                # ="=SI("+fxglo2 +" > 0;"+ " & "+fxglo2+" &  '%';"+fxglo2+"& " '%'")

                                fxglo2 = '(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*1/5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                    get_column_letter(j + 13)) + str(
                                    w)
                                fxglo = '=(' + str(get_column_letter(j + 16)) + str(w) + '*1+' + str(
                                    get_column_letter(j + 17)) + str(w) + '*0.5+' + str(
                                    get_column_letter(j + 18)) + str(w) + '*0)/' + str(
                                    get_column_letter(j + 13)) + str(
                                    w)  # pour  automatiser
                                # rapport.cell(w, j + 19).value = fglobal

                                # rapport.cell(w, j + 19).value =  fxglo
                                # rapport.cell(w, j + 19).value =  '=IF('+fxglo2 +' > 0;"+"'+ '& '+fxglo2+' & "%";'+fxglo2+' & "%")'
                                # pour marche mette ; ou lieu de  ,
                                rapport3.cell(w,
                                              j + 19).value = '=IF(' + fxglo2 + ' > 0," "' + '& ' + fxglo2 + ' & "%",' + fxglo2 + ' & "%")'
                                # rapport.cell(w, j + 19).value = 'IF(+'+str(fxglo2)+'  > 0,'+""+str(fxglo2)+"'%'"+',"False")'



                        ###############################################parite lie a les elemnt qui ont aps aucun titre rouge
                        else:

                            go = [1, 2, 3, 4, 5]
                            rapport3.cell(w, j + 1).value = 0
                            rapport3.cell(w, j + 6).value = "0%"
                            rapport3.cell(w, j + 1).border = thin_border
                            rapport3.cell(w, j + 6).border = thin_border
                            if s == len(go) - 1:
                                if a < len(list_indice3):
                                    rapport3.cell(4, j + 15).value = list_indice3[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    rapport3.cell(4, j - 3).value = list_indice3[
                                        a]  # mette le titre des partie des tableau  a partir de list list_indice
                                    a += 1
                                rapport3.cell(w, j + 8).border = thin_border
                                rapport3.cell(w, j + 9).border = thin_border
                                rapport3.cell(w - 6, j + 7).fill = PatternFill(patternType='solid',
                                                                               fgColor=self.my_gray)  # c'est pour siparation en gray '
                                rapport3.cell(w, j + 7).fill = PatternFill(patternType='solid', fgColor=self.my_gray)
                                rapport3.cell(w - 6, j + 21).fill = PatternFill(patternType='solid',
                                                                                fgColor=self.my_black)
                                rapport3.cell(w, j + 21).fill = PatternFill(patternType='solid',
                                                                            fgColor=self.my_black)
                                rapport3.cell(w, j + 8).value = 0  # N
                                rapport3.cell(w, j + 9).value = 0  # n
                                rapport3.cell(w, j + 10).value = '0%'  # KPI.1a Project Management
                                rapport3.cell(w, j + 11).value = '0%'
                                kpi_instance.append('0%')# pour l'envoyer
                                kpi_instance.append('0%')

                                rapport3.cell(w, j + 12).value = 0  # n-Nb"R"
                                rapport3.cell(w, j + 13).value = 0  # V
                                if go[2] != 4:
                                    rapport3.cell(w, j + 14).value = '0%'
                                else:
                                    rapport3.cell(w, j + 14).value = '0' + '%'

                self.kpi_raport2.append(kpi_instance)



        g_index = [i for i, x in enumerate(self.list_report3) if x == self.list_state_titles[
            0]]  # get l'index des case qui va prend la somme des couleur greende list fix
        o_index = [i for i, x in enumerate(self.list_report3) if x == self.list_state_titles[1]]
        r_index = [i for i, x in enumerate(self.list_report3) if x == self.list_state_titles[2]]
        na_index = [i for i, x in enumerate(self.list_report3) if x == self.list_state_titles[3]]
        ne_index = [i for i, x in enumerate(self.list_report3) if x == self.list_state_titles[4]]
        list_c = [g_index, o_index, r_index, na_index, ne_index]

        self.valeur = []
        self.list_c2 = []  # list des index des coloone de state des couleur

        for i in self.list_report3:

            for j in list_names_titles3:

                print(self.list_red,j)
                if i in j and i in self.list_red:

                    if list_names_titles3.index(j) not in self.valeur:

                        self.valeur.append(list_names_titles3.index(j))
                    for p in list_c:

                        self.z = []
                        if p == list_c[3] or p == list_c[4]:

                            for h in range(len(p)):
                                if p[h] not in self.z:
                                    self.z.append(p[h])
                            if self.z not in self.list_c2:
                                self.list_c2.append(self.z)
                        else:

                            for h in range(len(p)):
                                if h % 2 == 0:
                                    if p[h] not in self.z:
                                        self.z.append(p[h])
                            if self.z not in self.list_c2:
                                self.list_c2.append(self.z)

                        # for h in range(len(r_index)) :
                        # if h %2==0:
                        # if r_index[h] not in self.f:
                        #  self.f.append(r_index[h])

        w = 6
        traiter(self.list_c2)



